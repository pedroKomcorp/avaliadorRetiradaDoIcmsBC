import os

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image


def auto_adjust_column_widths(sheet):
    for column in sheet.columns:
        max_length = 0
        column = list(column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = adjusted_width


def create_spreadsheet(results, filename, client_name, client_cnpj):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Resultados"

# styles
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    fill_1 = PatternFill(start_color="e5d3c7", end_color="e5d3c7", fill_type="solid")
    fill_2 = PatternFill(start_color="cba797", end_color="cba797", fill_type="solid")

# header section
    header_fill = PatternFill(start_color="1F1E27", end_color="1F1E27", fill_type="solid")
    sheet.merge_cells('A1:H1')
    sheet['A1'] = f"CLIENTE: {client_name}"
    sheet['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    sheet['A1'].alignment = Alignment(horizontal="left", vertical="center")
    sheet['A1'].fill = header_fill
    sheet.row_dimensions[1].height = 35
    sheet.merge_cells('A2:H2')
    sheet['A2'] = f"CNPJ: {client_cnpj}"
    sheet['A2'].font = Font(bold=True, size=18, color="FFFFFF")
    sheet['A2'].alignment = Alignment(horizontal="left", vertical="center")
    sheet['A2'].fill = header_fill
    sheet['H1'].alignment = Alignment(vertical='bottom')
    sheet.row_dimensions[2].height = 35
    script_dir = os.path.dirname(os.path.abspath(__file__))
    img = Image(os.path.join(script_dir, "assets", "LOGOS_KOMBUSINESS_FUNDOESCURO_V2.ico"))
    img.width = 200.84
    img.height = 70.2
    img.anchor = 'G1'
    sheet.add_image(img)

# total section
    sheet.merge_cells('A4:B4')
    sheet['A4'] = 'Créditos gerados'
    sheet['A4'].font = Font(bold=True, size=14, color="FFFFFF")
    sheet['A4'].alignment = Alignment(horizontal="center", vertical="center")
    sheet['A4'].fill = header_fill
    sum_titles = ['Total PIS:', 'Total COFINS:', 'Total crédito:']
    empty_row_num = 0
    for i, title in enumerate(sum_titles, start=5):
        sheet[f'A{i}'] = title
        sheet[f'A{i}'].font = Font(bold=True, size=14)
        sheet[f'A{i}'].alignment = Alignment(horizontal="left", vertical="center")
        sheet[f'B{i}'].number_format = 'R$ #,##0.00'
        for col in 'AB':
            cell = sheet[f'{col}{i}']
            cell.fill = fill_1
            cell.border = thin_border

    # Adicionando uma linha de espaço


# pis section
    pis_headers = [
        "Competência", "Base de Cálculo PIS", "Valor PIS", "Base de Cálculo PIS sem ICMS", "Valor PIS sem ICMS",
        "Crédito PIS", "SELIC", "Crédito PIS Atualizado"
    ]
    start_row_pis_header = 9
    sheet.merge_cells(f'A{start_row_pis_header}:H{start_row_pis_header}')
    sheet[f'A{start_row_pis_header}'] = 'PIS'
    sheet[f'A{start_row_pis_header}'].font = Font(bold=True, size=14, color="FFFFFF")
    sheet[f'A{start_row_pis_header}'].alignment = Alignment(horizontal="center", vertical="center")
    sheet[f'A{start_row_pis_header}'].fill = header_fill
    start_row_pis = start_row_pis_header + 1
    for col_num, header in enumerate(pis_headers, 1):
        cell = sheet.cell(row=start_row_pis, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row_num, result in enumerate(results, start=start_row_pis + 1):
        pis_data = [
            result.comp, result.base_pis, result.valor_pis, result.base_pis_sem_icms, result.valor_pis_sem_icms,
            result.cred_pis, result.selic, result.cred_pis_atz
        ]
        fill = fill_1 if (row_num - start_row_pis) % 2 == 0 else fill_2
        for col_num, data in enumerate(pis_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = data
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if col_num == 7:  # Aplicar formato de porcentagem para a coluna SELIC
                cell.number_format = '0.00%'
            elif col_num != 1:  # Aplicar formato de moeda para valores monetários
                cell.number_format = 'R$ #,##0.00'
    total_row_num = start_row_pis + len(results) + 1
    total_label_cell = sheet.cell(row=total_row_num, column=7)
    total_label_cell.value = "TOTAL:"
    total_label_cell.fill = header_fill
    total_label_cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    sum_formula_pis = f"=SUM(H{start_row_pis + 1}:H{total_row_num - 1})"
    total_value_cell = sheet.cell(row=total_row_num, column=8)
    total_value_cell.value = sum_formula_pis
    total_value_cell.fill = header_fill
    total_value_cell.number_format = 'R$ #,##0.00'
    total_value_cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    sheet['B5'] = sum_formula_pis

# cofins section
    cofins_headers = [
        "Competência", "Base de Cálculo COFINS", "Valor COFINS", "Base de Cálculo COFINS sem ICMS",
        "Valor COFINS sem ICMS", "Crédito COFINS", "SELIC", "Crédito COFINS Atualizado"
    ]
    start_row_cofins_header = total_row_num + 2
    sheet.merge_cells(f'A{start_row_cofins_header}:H{start_row_cofins_header}')
    sheet[f'A{start_row_cofins_header}'] = 'COFINS'
    sheet[f'A{start_row_cofins_header}'].font = Font(bold=True, size=14, color="FFFFFF")
    sheet[f'A{start_row_cofins_header}'].alignment = Alignment(horizontal="center", vertical="center")
    sheet[f'A{start_row_cofins_header}'].fill = header_fill
    start_row_cofins = start_row_cofins_header + 1
    for col_num, header in enumerate(cofins_headers, 1):
        cell = sheet.cell(row=start_row_cofins, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row_num, result in enumerate(results, start=start_row_cofins + 1):
        cofins_data = [
            result.comp, result.base_cofins, result.valor_cofins, result.base_cofins_sem_icms,
            result.valor_cofins_sem_icms, result.cred_cofins, result.selic, result.cred_cofins_atz
        ]
        fill = fill_1 if (row_num - start_row_cofins) % 2 == 0 else fill_2
        for col_num, data in enumerate(cofins_data, 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell.value = data
            cell.fill = fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if col_num == 7:  # Aplicar formato de porcentagem para a coluna SELIC
                cell.number_format = '0.00%'
            elif col_num != 1:  # Aplicar formato de moeda para valores monetários
                cell.number_format = 'R$ #,##0.00'
    total_row_num_cofins = start_row_cofins + len(results) + 1
    total_label_cell_cofins = sheet.cell(row=total_row_num_cofins, column=7)
    total_label_cell_cofins.value = "TOTAL:"
    total_label_cell_cofins.fill = header_fill
    total_label_cell_cofins.font = Font(bold=True, color="FFFFFF")
    sum_formula_cofins = f"=SUM(H{start_row_cofins + 1}:H{total_row_num_cofins - 1})"
    total_value_cell_cofins = sheet.cell(row=total_row_num_cofins, column=8)
    total_value_cell_cofins.value = sum_formula_cofins
    total_value_cell_cofins.fill = header_fill
    total_value_cell_cofins.number_format = 'R$ #,##0.00'
    total_value_cell_cofins.font = Font(bold=True, color="FFFFFF")
    sheet['B6'] = sum_formula_cofins

    total_sum = f"=B5 + B6"
    sheet['B7'] = total_sum
    auto_adjust_column_widths(sheet)
    workbook.save(filename)
